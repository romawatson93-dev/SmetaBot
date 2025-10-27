import asyncio
import base64
import io
import os
import re
import logging
import subprocess
import shutil
import tempfile
import shutil
from threading import Lock
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from aiogram import Router, F
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import (
    Message,
    CallbackQuery,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    InputMediaPhoto,
)
from aiogram.types.input_file import BufferedInputFile
from aiogram.exceptions import TelegramBadRequest
from celery.exceptions import TimeoutError as CeleryTimeout

import bot.services.channels as channels_service

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.properties import PageSetupProperties

    _OPENPYXL_OK = True
except Exception:  # pragma: no cover - handled by runtime fallbacks
    openpyxl = None  # type: ignore
    get_column_letter = None  # type: ignore
    range_boundaries = None  # type: ignore
    PageMargins = None  # type: ignore
    PageSetupProperties = None  # type: ignore
    _OPENPYXL_OK = False
from bot.celery_client import get_celery
from bot.handlers.menu_common import (
    build_render_menu_keyboard,
    BTN_RENDER_PDF,
    BTN_RENDER_DOC,
    BTN_RENDER_XLSX,
    BTN_RENDER_PNG,
)

try:
    import fitz  # type: ignore

    _FITZ_OK = True
except Exception:
    _FITZ_OK = False

try:
    from PIL import Image, ImageDraw, ImageFont

    _PIL_OK = True
except Exception:
    _PIL_OK = False

from common.watermark import WATERMARK_SETTINGS, WatermarkSettings
from bot.storage import store_blob, load_blob, delete_blob, delete_many

router = Router()

MAX_FILE_SIZE = 20 * 1024 * 1024
PREVIEW_QUEUE_NAME = os.getenv("CELERY_PREVIEW_QUEUE", "preview")
PREVIEW_TASK_TIMEOUT = int(os.getenv("PREVIEW_TASK_TIMEOUT", "120"))
SOURCE_PREFIXES = {
    "pdf": "pdf",
    "docx": "doc",
    "xlsx": "xls",
    "png": "png",
}
logger = logging.getLogger(__name__)


def _format_error(prefix: str, exc: Exception, *, limit: int = 3500) -> str:
    message = str(exc)
    if len(message) > limit:
        message = message[:limit] + "…"
    return f"{prefix}: {message}"


async def _release_storage_for_items(items: List[Dict[str, Any]]) -> None:
    unique_keys: Set[str] = set()
    for item in items:
        source_key = item.get("source_key")
        if source_key:
            unique_keys.add(source_key)
        for page in item.get("pages") or []:
            for field in ("fullres_key", "source_key"):
                key = page.get(field)
                if key:
                    unique_keys.add(key)
    if unique_keys:
        await delete_many(unique_keys)

_RENDER_LOCKS: Dict[int, asyncio.Lock] = {}
_USE_CELERY_PUBLISH = os.getenv("ENABLE_CELERY_PUBLISH", "1").lower() not in {"0", "false", "no"}
_WM_TILE_CACHE_SIZE = 32
_WM_TILE_CACHE: Dict[Tuple[str, Tuple[str, str], int, int, int, Tuple[int, int, int], int, int, int], Image.Image] = {}
_WM_TILE_LOCK = Lock()


def _get_render_lock(user_id: int) -> asyncio.Lock:
    lock = _RENDER_LOCKS.get(user_id)
    if lock is None:
        lock = asyncio.Lock()
        _RENDER_LOCKS[user_id] = lock
    return lock


async def _ensure_page_original_bytes(page: Dict[str, Any]) -> Optional[bytes]:
    cached = page.get("original_bytes")
    if cached:
        return cached
    fullres_key = page.get("fullres_key")
    if not fullres_key:
        return None
    try:
        payload = await load_blob(fullres_key, delete=False)
    except Exception as exc:
        logger.warning("render: failed to load fullres page (key=%s): %s", fullres_key, exc)
        return None
    page["original_bytes"] = payload
    return payload


async def _fetch_preview_from_worker(
    render_format: str,
    filename: str,
    *,
    storage_key: str | None,
    blob: bytes | None,
) -> Dict[str, Any]:
    celery_app = get_celery()
    queue_name = PREVIEW_QUEUE_NAME
    kwargs = {
        "filename": filename,
        "render_format": render_format,
    }
    if storage_key:
        kwargs["file_key"] = storage_key
    elif blob is not None:
        kwargs["file_b64"] = base64.b64encode(blob).decode("ascii")
    else:
        raise RuntimeError("Preview worker requires either storage key or raw payload.")

    async_result = celery_app.send_task(
        "tasks.preview.generate_preview_task",
        kwargs=kwargs,
        queue=queue_name,
    )
    try:
        result = await asyncio.to_thread(async_result.get, timeout=PREVIEW_TASK_TIMEOUT)
        if not isinstance(result, dict):
            raise RuntimeError("Неверный ответ превью-задачи.")
        return result
    except CeleryTimeout as exc:
        raise RuntimeError("Превью готовится дольше обычного. Попробуйте повторить позже.") from exc
    except Exception as exc:
        raise RuntimeError(str(exc)) from exc
    finally:
        try:
            async_result.forget()
        except Exception:
            pass

ROW_GAP_TOLERANCE = 3


class RenderSession(StatesGroup):
    waiting_file = State()
    idle = State()
    waiting_wm_text = State()


def _sanitize_basename(name: str) -> str:
    base = os.path.splitext(name)[0] or "document"
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", base)
    return base[:48] or "document"


def _convert_pdf(data: bytes, filename: str) -> List[Dict[str, Any]]:
    if not _FITZ_OK:
        raise RuntimeError("PyMuPDF (fitz) недоступен в окружении.")
    doc = fitz.open(stream=data, filetype="pdf")
    pages: List[Dict[str, Any]] = []
    try:
        total = doc.page_count
        base = Path(filename).stem or "page"
        for idx, page in enumerate(doc, start=1):
            pix = page.get_pixmap(dpi=300, alpha=False)
            out_name = f"{base}-{idx:02}.png" if total > 1 else f"{base}.png"
            pages.append(
                {
                    "filename": out_name,
                    "content": pix.tobytes("png"),
                    "page_index": idx,
                    "pages_total": total,
                }
            )
    finally:
        doc.close()
    return pages



def _wrap_png_as_pages(png_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    base_name = Path(filename).stem or "image"
    return [{
        "filename": f"{base_name}.png",
        "content": png_bytes,
        "page_index": 1,
        "pages_total": 1,
    }]

def _convert_doc_to_pdf_bytes(doc_bytes: bytes, suffix: str) -> bytes:
    print(f"[render] _convert_doc_to_pdf_bytes start suffix={suffix} size={len(doc_bytes)}", flush=True)
    suffix = suffix.lower()
    if suffix not in {".doc", ".docx"}:
        raise ValueError("Поддерживаются только файлы DOC и DOCX.")
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            src_path = tmpdir_path / f"source{suffix.lower()}"
            src_path.write_bytes(doc_bytes)
            binary = shutil.which("libreoffice") or shutil.which("soffice")
            if not binary:
                raise RuntimeError("LibreOffice не установлен в среде исполнения.")
            cmd = [
                binary,
                "--headless",
                "--nologo",
                "--nodefault",
                "--nofirststartwizard",
                "--norestore",
                "--nolockcheck",
                "--convert-to",
                "pdf",
                "--outdir",
                str(tmpdir_path),
                str(src_path),
            ]
            env = os.environ.copy()
            env.setdefault("HOME", str(tmpdir_path))
            env.setdefault("TMPDIR", str(tmpdir_path))
            proc = subprocess.run(
                cmd,
                capture_output=True,
                timeout=240,
                cwd=tmpdir_path,
                env=env,
            )
            if proc.returncode != 0:
                raise RuntimeError(
                    f"LibreOffice не смог конвертировать файл: {proc.stderr.decode(errors='ignore') or proc.stdout.decode(errors='ignore')}"
                )
            pdf_path = src_path.with_suffix(".pdf")
            if not pdf_path.exists():
                candidates = list(tmpdir_path.glob("*.pdf"))
                if not candidates:
                    raise RuntimeError("LibreOffice не создал PDF-файл на выходе.")
                pdf_path = candidates[0]
            result = pdf_path.read_bytes()
            print(f"[render] _convert_doc_to_pdf_bytes done output={len(result)}", flush=True)
            return result
    except FileNotFoundError as e:
        raise RuntimeError("LibreOffice не установлен в среде исполнения.") from e
    except subprocess.TimeoutExpired as e:
        raise RuntimeError("Конвертация LibreOffice заняла слишком много времени и была остановлена.") from e

def _convert_doc_to_png_bytes(doc_bytes: bytes, suffix: str, filename: str) -> List[Dict[str, Any]]:
    suffix = suffix.lower()
    if suffix not in {".doc", ".docx"}:
        raise ValueError("Only DOC and DOCX files are supported.")

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        src_path = tmpdir_path / f"source{suffix}"
        src_path.write_bytes(doc_bytes)
        pdf_path = tmpdir_path / "export.pdf"
        png_pattern = tmpdir_path / "page-%03d.png"

        binary = shutil.which("libreoffice") or shutil.which("soffice")
        if not binary:
            raise RuntimeError("LibreOffice binary is not available.")

        safe_home = Path(tempfile.mkdtemp(prefix="lo_home_"))
        try:
            env = os.environ.copy()
            env["HOME"] = str(safe_home)
            env["TMPDIR"] = str(tmpdir_path)
            env["SAL_USE_VCLPLUGIN"] = "headless"

            pdf_cmd = [
                binary,
                "--headless",
                "--nologo",
                "--nodefault",
                "--nofirststartwizard",
                "--norestore",
                "--convert-to",
                "pdf:writer_pdf_Export:EmbedStandardFonts=true;UseTaggedPDF=false;UseLosslessCompression=true;ExportNotes=false;SkipEmptyPages=false;ExportBookmarks=false;SelectPdfVersion=1",
                str(src_path.resolve()),
                "--outdir",
                str(tmpdir_path),
            ]
            proc_pdf = subprocess.run(
                pdf_cmd,
                cwd=tmpdir_path,
                env=env,
                capture_output=True,
                text=True,
                timeout=240,
            )
            if proc_pdf.returncode != 0:
                raise RuntimeError(
                    "LibreOffice failed to export PDF.\n"
                    + "Command: " + " ".join(pdf_cmd) + "\n"
                    + f"STDOUT:\n{proc_pdf.stdout}\nSTDERR:\n{proc_pdf.stderr}"
                )

            if not pdf_path.exists():
                candidates = sorted(tmpdir_path.glob("*.pdf"))
                if not candidates:
                    raise RuntimeError(
                        "LibreOffice did not produce PDF output.\n"
                        + "Command: " + " ".join(pdf_cmd) + "\n"
                        + f"STDOUT:\n{proc_pdf.stdout}\nSTDERR:\n{proc_pdf.stderr}"
                    )
                pdf_path = candidates[0]
            print(
                f"[render] LibreOffice output PDF: {pdf_path.name} ({pdf_path.stat().st_size} bytes)",
                flush=True,
            )

            gs_cmd = [
                "gs",
                "-dSAFER",
                "-dBATCH",
                "-dNOPAUSE",
                "-dQUIET",
                "-sDEVICE=png16m",
                "-r300",
                f"-sOutputFile={png_pattern}",
                str(pdf_path),
            ]
            proc_gs = subprocess.run(
                gs_cmd,
                cwd=tmpdir_path,
                capture_output=True,
                text=True,
                timeout=240,
            )
            if proc_gs.returncode != 0:
                raise RuntimeError(
                    "Ghostscript failed to render PDF.\n"
                    + "Command: " + " ".join(gs_cmd) + "\n"
                    + f"STDOUT:\n{proc_gs.stdout}\nSTDERR:\n{proc_gs.stderr}"
                )

            png_paths = sorted(tmpdir_path.glob("page-*.png"))
            if not png_paths:
                raise RuntimeError(
                    "Ghostscript did not produce PNG files.\n"
                    + "Command: " + " ".join(gs_cmd) + "\n"
                    + f"STDOUT:\n{proc_gs.stdout}\nSTDERR:\n{proc_gs.stderr}"
                )

            base_name = Path(filename).stem or "page"
            pages: List[Dict[str, Any]] = []
            total = len(png_paths)
            for idx, path in enumerate(png_paths, start=1):
                out_name = f"{base_name}-{idx:03}.png" if total > 1 else f"{base_name}.png"
                pages.append({"filename": out_name, "content": path.read_bytes(), "page_index": idx, "pages_total": total})
            return pages
        finally:
            shutil.rmtree(safe_home, ignore_errors=True)
































def _libreoffice_convert_bytes(src_bytes: bytes, suffix: str, target: str) -> bytes:
    suffix = suffix.lower()
    binary = shutil.which("libreoffice") or shutil.which("soffice")
    if not binary:
        raise RuntimeError("LibreOffice не обнаружен в PATH.")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        src_path = tmpdir_path / f"source{suffix}"
        src_path.write_bytes(src_bytes)
        cmd = [
            binary,
            "--headless",
            "--nologo",
            "--nodefault",
            "--nofirststartwizard",
            "--norestore",
            "--nolockcheck",
            "--convert-to",
            target,
            "--outdir",
            str(tmpdir_path),
            str(src_path),
        ]
        env = os.environ.copy()
        env.setdefault("HOME", str(tmpdir_path))
        env.setdefault("TMPDIR", str(tmpdir_path))
        proc = subprocess.run(
            cmd,
            capture_output=True,
            timeout=240,
            cwd=tmpdir_path,
            env=env,
        )
        if proc.returncode != 0:
            stderr = proc.stderr.decode(errors="ignore") or proc.stdout.decode(errors="ignore")
            raise RuntimeError(f"LibreOffice не смог конвертировать {suffix} в {target}: {stderr}")
        target_suffix = target.split(":", 1)[0]
        if not target_suffix.startswith("."):
            target_suffix = "." + target_suffix
        out_path = src_path.with_suffix(target_suffix)
        if not out_path.exists():
            candidates = list(tmpdir_path.glob(f"*{target_suffix}"))
            if not candidates:
                raise RuntimeError(f"LibreOffice не создал файл {target_suffix}.")
            out_path = candidates[0]
        return out_path.read_bytes()


def _convert_excel_to_pdf_bytes(excel_bytes: bytes, suffix: str) -> bytes:
    return _libreoffice_convert_bytes(excel_bytes, suffix, "pdf:calc_pdf_Export")


def _prepare_excel_bytes_for_openpyxl(excel_bytes: bytes, suffix: str) -> tuple[bytes, str]:
    suffix = suffix.lower() or ".xlsx"
    if suffix in {".xlsx", ".xlsm"}:
        return excel_bytes, suffix
    if suffix in {".xls", ".ods", ".fods"}:
        converted = _libreoffice_convert_bytes(excel_bytes, suffix, "xlsx")
        return converted, ".xlsx"
    raise ValueError("Поддерживаются только Excel/ODS файлы.")


def _cell_has_value(cell) -> bool:
    value = cell.value
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _worksheet_detect_tables(ws) -> List[tuple[int, int, int, int]]:
    if range_boundaries is None:
        return []
    try:
        min_col, min_row, max_col, max_row = range_boundaries(ws.calculate_dimension())
    except ValueError:
        return []

    row_to_cols: Dict[int, Set[int]] = {}

    def _touch_cell(row_idx: int, col_idx: int) -> None:
        row_to_cols.setdefault(row_idx, set()).add(col_idx)

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if _cell_has_value(cell):
                _touch_cell(cell.row, cell.column)

    for img in getattr(ws, "_images", []):
        anchor = getattr(img, "anchor", None)
        if anchor is None:
            continue
        if hasattr(anchor, "_from"):
            cell_from = anchor._from
            row_idx = getattr(cell_from, "row", None)
            col_idx = getattr(cell_from, "col", None)
            if row_idx is not None and col_idx is not None:
                _touch_cell(int(row_idx) + 1, int(col_idx) + 1)
        elif hasattr(anchor, "row") and hasattr(anchor, "col"):
            _touch_cell(int(anchor.row) + 1, int(anchor.col) + 1)

    if not row_to_cols:
        return []

    sorted_rows = sorted(row_to_cols.keys())
    row_groups: List[List[int]] = []
    current_group: List[int] = []
    for row_idx in sorted_rows:
        if not current_group:
            current_group = [row_idx]
            continue
        if row_idx - current_group[-1] <= ROW_GAP_TOLERANCE + 1:
            current_group.append(row_idx)
        else:
            row_groups.append(current_group)
            current_group = [row_idx]
    if current_group:
        row_groups.append(current_group)

    regions: List[tuple[int, int, int, int]] = []
    for rows in row_groups:
        columns = sorted({col for r in rows for col in row_to_cols.get(r, ())})
        if not columns:
            continue
        col_groups: List[List[int]] = []
        current_cols: List[int] = []
        for col_idx in columns:
            if not current_cols:
                current_cols = [col_idx]
                continue
            if col_idx - current_cols[-1] <= 1:
                current_cols.append(col_idx)
            else:
                col_groups.append(current_cols)
                current_cols = [col_idx]
        if current_cols:
            col_groups.append(current_cols)

        for cols in col_groups:
            min_row_group = rows[0]
            max_row_group = rows[-1]
            min_col_group = cols[0]
            max_col_group = cols[-1]

            # expand upward/downward if there is nearby content (e.g., headers, footers)
            for offset in range(1, ROW_GAP_TOLERANCE + 2):
                candidate = min_row_group - offset
                if candidate < 1:
                    break
                if row_to_cols.get(candidate):
                    min_row_group = candidate
                else:
                    break
            for offset in range(1, ROW_GAP_TOLERANCE + 1):
                candidate = max_row_group + offset
                if candidate > max_row:
                    break
                if row_to_cols.get(candidate):
                    max_row_group = candidate
                else:
                    break

            data_cells = sum(
                1
                for r in range(min_row_group, max_row_group + 1)
                for c in range(min_col_group, max_col_group + 1)
                if c in row_to_cols.get(r, ())
            )
            area = (max_row_group - min_row_group + 1) * (max_col_group - min_col_group + 1)
            if data_cells == 0:
                continue
            if data_cells < 3 and area <= 3:
                continue
            regions.append((min_row_group, max_row_group, min_col_group, max_col_group))

    regions.sort(key=lambda b: (b[0], b[2]))
    return regions


def _range_to_a1(bounds: tuple[int, int, int, int]) -> str:
    min_row, max_row, min_col, max_col = bounds
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return f"{start}:{end}"


def _export_excel_region_to_pdf(excel_bytes: bytes, suffix: str, sheet_name: str, bounds: tuple[int, int, int, int]) -> bytes:
    if not _OPENPYXL_OK:
        raise RuntimeError('openpyxl не установлен, экспорт Excel недоступен.')
    suffix = (suffix or '.xlsx').lower()
    if suffix not in {'.xlsx', '.xlsm', '.xls', '.ods', '.fods'}:
        suffix = '.xlsx'

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_path = tmp_path / f'source{suffix}'
        src_path.write_bytes(excel_bytes)

        load_kwargs = {'data_only': False}
        if suffix == '.xlsm':
            load_kwargs['keep_vba'] = True
        wb = openpyxl.load_workbook(src_path, **load_kwargs)
        try:
            if sheet_name not in wb.sheetnames:
                raise RuntimeError(f'Лист {sheet_name!r} не найден в книге.')
            for ws in wb.worksheets:
                ws.sheet_state = 'hidden' if ws.title != sheet_name else 'visible'
            ws = wb[sheet_name]
            area = _range_to_a1(bounds)
            ws.print_area = area
            if PageSetupProperties is not None:
                ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            ws.page_setup.fitToWidth = 1  # type: ignore[attr-defined]
            ws.page_setup.fitToHeight = 1  # type: ignore[attr-defined]
            orientation = 'landscape' if (bounds[3] - bounds[2]) > (bounds[1] - bounds[0]) else 'portrait'
            ws.page_setup.orientation = orientation  # type: ignore[attr-defined]
            if PageMargins is not None:
                ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.3, bottom=0.3, header=0.1, footer=0.1)
            wb.active = wb.sheetnames.index(sheet_name)
            wb.save(src_path)
        finally:
            wb.close()

        modified_bytes = src_path.read_bytes()

    return _convert_excel_to_pdf_bytes(modified_bytes, suffix)


def _extract_excel_tables(excel_bytes: bytes, filename: str) -> dict[str, Any]:
    if not _OPENPYXL_OK:
        raise RuntimeError("openpyxl не установлен, анализ Excel недоступен.")
    suffix = Path(filename).suffix or ".xlsx"
    prepared_bytes, prepared_suffix = _prepare_excel_bytes_for_openpyxl(excel_bytes, suffix)
    wb = openpyxl.load_workbook(io.BytesIO(prepared_bytes), data_only=True)
    try:
        total_sheets = len(wb.worksheets)
        orig_base = (os.path.splitext(filename)[0] or "document").strip()
        base_name = _sanitize_basename(filename)
        tables: list[dict[str, Any]] = []
        for sheet_index, ws in enumerate(wb.worksheets):
            regions = _worksheet_detect_tables(ws)
            if not regions:
                continue
            tables_in_sheet = len(regions)
            for table_idx, bounds in enumerate(regions, start=1):
                pdf_bytes = _export_excel_region_to_pdf(prepared_bytes, prepared_suffix, ws.title, bounds)
                png_pages = _convert_pdf(
                    pdf_bytes,
                    f"{base_name}-sheet{sheet_index + 1}-tbl{table_idx}.pdf",
                )
                if not png_pages:
                    continue
                first = png_pages[0]
                display_name = (orig_base or "document") + ".png"
                tables.append(
                    {
                        "filename": display_name,
                        "content": first["content"],
                        "sheet_name": ws.title,
                        "table_range": _range_to_a1(bounds),
                        "sheet_index": sheet_index,
                        "sheets_total": total_sheets,
                        "table_index": table_idx,
                        "tables_in_sheet": tables_in_sheet,
                        "base_name": base_name,
                        "display_name": display_name,
                    }
                )
        return {"pages": tables, "sheets_total": total_sheets}
    finally:
        wb.close()


def _load_font(size: int, settings: WatermarkSettings = WATERMARK_SETTINGS) -> ImageFont.FreeTypeFont:
    candidates = [
        settings.font_preferred,
        "Roboto-Bold.ttf",
        "Roboto-Regular.ttf",
        "/usr/share/fonts/truetype/roboto/Roboto-Bold.ttf",
        "/usr/share/fonts/truetype/roboto/Roboto-Regular.ttf",
        settings.font_fallback,
        "DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        try:
            return ImageFont.truetype(path, size=size)
        except Exception:
            continue
    return ImageFont.load_default()


def _wm_cache_key(
    text: str,
    font: ImageFont.FreeTypeFont,
    cfg: WatermarkSettings,
    tile_w: int,
    tile_h: int,
) -> Tuple[str, Tuple[str, str], int, int, int, Tuple[int, int, int], int, int, int]:
    font_name = font.getname()
    font_size = getattr(font, "size", 0)
    return (
        text,
        font_name,
        font_size,
        tile_w,
        tile_h,
        cfg.color,
        cfg.opacity,
        cfg.angle,
        cfg.text_offset,
    )


def _wm_get_tile(
    text: str,
    font: ImageFont.FreeTypeFont,
    cfg: WatermarkSettings,
    tile_w: int,
    tile_h: int,
) -> Image.Image:
    key = _wm_cache_key(text, font, cfg, tile_w, tile_h)
    with _WM_TILE_LOCK:
        cached = _WM_TILE_CACHE.get(key)
        if cached is not None:
            return cached

    tile = Image.new("RGBA", (tile_w, tile_h), (0, 0, 0, 0))
    drawer = ImageDraw.Draw(tile)
    bbox = drawer.textbbox((0, 0), text, font=font)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    color = (*cfg.color, max(16, min(255, cfg.opacity)))
    if cfg.text_offset < 0:
        pos_x = (tile_w - tw) // 2
        pos_y = (tile_h - th) // 2
    else:
        pos_x = cfg.text_offset
        pos_y = cfg.text_offset
    drawer.text((pos_x, pos_y), text, font=font, fill=color)
    rotated = tile.rotate(cfg.angle, expand=True)

    with _WM_TILE_LOCK:
        if len(_WM_TILE_CACHE) >= _WM_TILE_CACHE_SIZE:
            _WM_TILE_CACHE.pop(next(iter(_WM_TILE_CACHE)))
        _WM_TILE_CACHE[key] = rotated
    return rotated


def _watermark_bytes(png_bytes: bytes, text: str) -> bytes:
    if not _PIL_OK:
        raise RuntimeError("Pillow недоступен для нанесения водяного знака.")
    cfg = WATERMARK_SETTINGS
    with Image.open(io.BytesIO(png_bytes)).convert("RGBA") as img:
        width, height = img.size
        overlay = Image.new("RGBA", (width, height), (0, 0, 0, 0))
        font_size = max(cfg.min_font_size, int(max(width, height) * cfg.font_scale))
        font = _load_font(font_size, cfg)

        tile_w = max(64, int(width * cfg.tile_scale_x))
        tile_h = max(64, int(height * cfg.tile_scale_y))
        rotated = _wm_get_tile(text, font, cfg, tile_w, tile_h)

        base_step = max(1, cfg.step)
        step_x = max(base_step, rotated.width // 2)
        step_y = max(base_step, rotated.height // 2)
        for x in range(-rotated.width, width + rotated.width, step_x):
            for y in range(-rotated.height, height + rotated.height, step_y):
                overlay.alpha_composite(rotated, dest=(x, y))
        stamped = Image.alpha_composite(img, overlay).convert("RGB")
    out = io.BytesIO()
    stamped.save(out, format="PNG")
    return out.getvalue()


async def _apply_watermark_to_items(items: List[Dict[str, Any]], text: str) -> None:
    if not _PIL_OK:
        raise RuntimeError("Функция водяного знака недоступна (Pillow не установлен).")

    for item in items:
        for page in item["pages"]:
            await _ensure_page_original_bytes(page)

    loop = asyncio.get_running_loop()

    def _work() -> None:
        for item in items:
            for page in item["pages"]:
                source = page.get("original_bytes") or page.get("preview_original_bytes")
                if not source:
                    continue
                page["watermarked_bytes"] = _watermark_bytes(source, text)
                page["preview_watermarked_bytes"] = None

    await loop.run_in_executor(None, _work)


async def _ensure_watermark_for_all(items: List[Dict[str, Any]], text: str) -> None:
    pending: List[Dict[str, Any]] = []
    for item in items:
        need = any(page.get("watermarked_bytes") is None for page in item["pages"])
        if need:
            pending.append(item)
    if pending:
        await _apply_watermark_to_items(pending, text)


def _clear_watermarks(items: List[Dict[str, Any]]) -> None:
    for item in items:
        for page in item["pages"]:
            page["watermarked_bytes"] = None
            page["preview_watermarked_bytes"] = None


def _ensure_preview_bytes(page: Dict[str, Any], watermarked: bool) -> bytes | None:
    if watermarked:
        cached = page.get("preview_watermarked_bytes")
        source = page.get("watermarked_bytes")
        target_key = "preview_watermarked_bytes"
    else:
        cached = page.get("preview_original_bytes")
        source = page.get("original_bytes") or page.get("preview_original_bytes")
        target_key = "preview_original_bytes"

    if cached:
        return cached
    if source is None:
        return None
    if not _PIL_OK:
        return source
    try:
        with Image.open(io.BytesIO(source)) as img:
            img = img.convert("RGB")
            img.thumbnail((1600, 1600), Image.LANCZOS)
            out = io.BytesIO()
            img.save(out, format="JPEG", quality=85, optimize=True)
            preview = out.getvalue()
            page[target_key] = preview
            return preview
    except Exception:
        return source


def _flatten_pages(items: List[Dict[str, Any]]) -> List[Tuple[int, int]]:
    order: List[Tuple[int, int]] = []
    for item_idx, item in enumerate(items):
        for page_idx, _ in enumerate(item["pages"]):
            order.append((item_idx, page_idx))
    return order


def _card_keyboard(
    total_pages: int,
    watermark_active: bool,
    allow_navigation: bool,
    render_format: str,
    current_selected: bool,
    selection_enabled: bool,
) -> InlineKeyboardMarkup:
    rows: List[List[InlineKeyboardButton]] = []
    if allow_navigation:
        rows.append([InlineKeyboardButton(text="◀️ Предыдущий", callback_data="render:prev")])
        rows.append([InlineKeyboardButton(text="▶️ Следующий", callback_data="render:next")])

    if selection_enabled:
        if render_format == "xlsx":
            select_text = "☑️ Убрать из выбора" if current_selected else "✅ Выбрать таблицу"
        elif render_format == "docx":
            select_text = "☑️ Убрать страницу" if current_selected else "✅ Выбрать страницу"
        else:
            select_text = "☑️ Убрать страницу" if current_selected else "✅ Добавить страницу"
        rows.append([InlineKeyboardButton(text=select_text, callback_data="render:toggle")])

    if watermark_active:
        rows.append([InlineKeyboardButton(text="🚫 Убрать водяной знак", callback_data="render:wm:clear")])
    else:
        rows.append([InlineKeyboardButton(text="🖋️ Водяной знак", callback_data="render:wm:set")])
    rows.append([InlineKeyboardButton(text="➕ Добавить файл", callback_data="render:add")])
    upload_text = "▶️ Продолжить" if render_format in ("xlsx", "docx") else "📤 Загрузить в канал"
    rows.append([InlineKeyboardButton(text=upload_text, callback_data="render:upload")])
    rows.append([InlineKeyboardButton(text="✖️ Отмена", callback_data="render:cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _build_caption(
    items: List[Dict[str, Any]],
    flat_index: int,
    total_pages: int,
    wm_text: str | None,
    render_format: str,
) -> str:
    flat = _flatten_pages(items)
    item_idx, page_idx = flat[flat_index]
    item = items[item_idx]
    page = item["pages"][page_idx]

    source_name = item.get("source") or page.get("filename") or "document"
    lines = [f"Страница {flat_index + 1} из {total_pages}", f"Файл: {source_name}"]

    output_name = page.get("filename")
    if output_name and output_name != source_name:
        lines.append(f"PNG: {output_name}")

    if render_format == "xlsx":
        sheet_total = page.get("sheets_total")
        sheet_index = page.get("sheet_index")
        sheet_name = page.get("sheet_name")
        tables_total = page.get("tables_in_sheet")
        table_index = page.get("table_index")
        if sheet_name:
            if sheet_index is not None and sheet_total:
                lines.append(f"Лист: {sheet_name} ({sheet_index + 1}/{sheet_total})")
            else:
                lines.append(f"Лист: {sheet_name}")
        elif sheet_index is not None and sheet_total:
            lines.append(f"Лист: {sheet_index + 1}/{sheet_total}")
        if sheet_total and not any(line.startswith("Листов в файле") for line in lines):
            lines.append(f"Листов в файле: {sheet_total}")
        if table_index:
            if tables_total:
                lines.append(f"Таблица: {table_index}/{tables_total}")
            else:
                lines.append(f"Таблица № {table_index}")
        table_range = page.get("table_range")
        if table_range:
            lines.append(f"Область: {table_range}")
        selected = page.get("selected", True)
        lines.append(f"Статус: {'✅ в подборке' if selected else '▫️ пропущено'}")
    elif render_format in {"docx", "pdf", "png"}:
        selected = page.get("selected", True)
        page_index = page.get("page_index")
        pages_total = page.get("pages_total")
        if render_format != "png":
            label = "Страница Word" if render_format == "docx" else "Страница PDF"
            if page_index and pages_total:
                lines.append(f"{label}: {page_index}/{pages_total}")
            elif page_index:
                lines.append(f"{label}: {page_index}")
        lines.append(f"Статус: {'✅ в подборке' if selected else '▫️ пропущено'}")

    if wm_text:
        lines.append(f"Водяной знак: «{wm_text}»")
    lines.append("")
    lines.append("Список файлов:")
    for item in items:
        for entry in item["pages"]:
            if render_format == "xlsx":
                prefix = "✅" if entry.get("selected", True) else "▫️"
                details: List[str] = []
                table_index = entry.get("table_index")
                tables_total = entry.get("tables_in_sheet")
                if table_index:
                    if tables_total:
                        details.append(f"табл. {table_index}/{tables_total}")
                    else:
                        details.append(f"табл. {table_index}")
                sheet_name = entry.get("sheet_name")
                if sheet_name:
                    details.append(str(sheet_name))
                table_range = entry.get("table_range")
                if table_range:
                    details.append(str(table_range))
                suffix = f" ({' — '.join(details)})" if details else ""
                lines.append(f"• {prefix} {entry['filename']}{suffix}")
            elif render_format in {"docx", "pdf", "png"}:
                prefix = "✅" if entry.get("selected", True) else "▫️"
                details: List[str] = []
                page_index = entry.get("page_index")
                pages_total = entry.get("pages_total")
                if render_format != "png":
                    if page_index and pages_total:
                        details.append(f"стр. {page_index}/{pages_total}")
                    elif page_index:
                        details.append(f"стр. {page_index}")
                suffix = f" ({' — '.join(details)})" if details else ""
                lines.append(f"• {prefix} {entry['filename']}{suffix}")
            else:
                lines.append(f"• {entry['filename']}")
    return "\n".join(lines)


async def _update_render_card(bot, chat_id: int, state: FSMContext, *, focus: str | None = None) -> None:
    data = await state.get_data()
    items: List[Dict[str, Any]] = list(data.get("render_items") or [])
    if not items:
        return
    render_format = (data.get("render_format") or "pdf").lower()
    flat = _flatten_pages(items)
    if not flat:
        return
    total = len(flat)
    index = data.get("render_index", 0)
    if focus == "last":
        index = total - 1
    elif focus == "first":
        index = 0
    index = max(0, min(index, total - 1))
    await state.update_data(render_index=index)

    item_idx, page_idx = flat[index]
    page = items[item_idx]["pages"][page_idx]
    wm_text: str | None = data.get("render_wm_text")
    if wm_text:
        await _ensure_watermark_for_all([items[item_idx]], wm_text)
        preview_bytes = _ensure_preview_bytes(page, True)
    else:
        preview_bytes = _ensure_preview_bytes(page, False)

    caption = _build_caption(items, index, total, wm_text, render_format)
    selection_enabled = render_format in {"xlsx", "docx", "pdf", "png"}
    current_selected = page.get("selected", True)
    keyboard = _card_keyboard(
        total,
        watermark_active=bool(wm_text),
        allow_navigation=total > 1,
        render_format=render_format,
        current_selected=current_selected,
        selection_enabled=selection_enabled,
    )
    card_mid = data.get("render_card_mid")

    if preview_bytes is None:
        if card_mid:
            try:
                await bot.delete_message(chat_id, card_mid)
            except Exception:
                pass
        sent = await bot.send_message(chat_id, caption, reply_markup=keyboard)
        await state.update_data(render_card_mid=sent.message_id)
        return

    media = InputMediaPhoto(media=BufferedInputFile(preview_bytes, filename="preview.jpg"), caption=caption)

    if card_mid:
        try:
            await bot.edit_message_media(
                chat_id=chat_id,
                message_id=card_mid,
                media=media,
                reply_markup=keyboard,
            )
            await state.update_data(render_card_mid=card_mid)
            return
        except TelegramBadRequest:
            # fall back to re-sending
            pass
        except Exception:
            try:
                await bot.delete_message(chat_id, card_mid)
            except Exception:
                pass

    try:
        sent = await bot.send_photo(
            chat_id=chat_id,
            photo=BufferedInputFile(preview_bytes, filename="preview.jpg"),
            caption=caption,
            reply_markup=keyboard,
        )
    except TelegramBadRequest:
        sent = await bot.send_message(chat_id, caption, reply_markup=keyboard)
    await state.update_data(render_card_mid=sent.message_id)


async def _clear_render_context(bot, chat_id: int, state: FSMContext) -> None:
    data = await state.get_data()
    card_mid = data.get("render_card_mid")
    choose_mid = data.get("render_choose_mid")
    items = list(data.get("render_items") or [])
    if card_mid:
        try:
            await bot.delete_message(chat_id, card_mid)
        except Exception:
            pass
    if choose_mid:
        try:
            await bot.delete_message(chat_id, choose_mid)
        except Exception:
            pass
    if items:
        await _release_storage_for_items(items)
    for key in ("render_items", "render_card_mid", "render_choose_mid", "render_channels", "render_index", "render_wm_text"):
        data.pop(key, None)
    await state.set_data(data)
    await state.set_state(None)


async def reset_render_state(state: FSMContext) -> None:
    data = await state.get_data()
    for key in ("render_items", "render_card_mid", "render_choose_mid", "render_channels", "render_index", "render_wm_text"):
        data.pop(key, None)
    await state.set_data(data)
    await state.set_state(None)


async def _fetch_recent_channels(contractor_id: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    try:
        contractor_id_int = int(contractor_id)
    except ValueError:
        contractor_id_int = int(contractor_id or 0)
    channels = await channels_service.list_channels(contractor_id_int, limit=5)
    for item in channels:
        # Используем tg_chat_id для публикации в воркере
        chat_id = int(item.get("tg_chat_id") or item.get("channel_id", 0))
        rows.append({"channel_id": chat_id, "title": item["title"]})
    return rows


@router.message(F.text == BTN_RENDER_PDF)
async def render_pdf_start(m: Message, state: FSMContext):
    if not _FITZ_OK:
        await m.answer("Конвертация PDF недоступна: библиотека PyMuPDF не установлена.")
        return
    logger.info("render: start PDF session user=%s", m.from_user.id)
    await reset_render_state(state)
    await state.set_state(RenderSession.waiting_file)
    await state.update_data(
        render_format="pdf",
        render_items=[],
        render_card_mid=None,
        render_choose_mid=None,
        render_channels={},
        render_index=0,
        render_wm_text=None,
    )
    await m.answer(
        "Пришлите PDF-файл как документ — мы конвертируем его в PNG 300 DPI. "
        "⚠️ Ограничение Telegram: размер файла до 20 МБ."
    )


@router.message(F.text == BTN_RENDER_PNG)
async def render_png_start(m: Message, state: FSMContext):
    logger.info("render: start PNG session user=%s", m.from_user.id)
    await reset_render_state(state)
    await state.set_state(RenderSession.waiting_file)
    await state.update_data(
        render_format="png",
        render_items=[],
        render_card_mid=None,
        render_choose_mid=None,
        render_channels={},
        render_index=0,
        render_wm_text=None,
    )
    await m.answer("Пришлите PNG-файл как документ (до 20 МБ). Эти файлы мы можем сразу отправить в канал, добавить водяной знак и выбрать нужные страницы.")


@router.message(F.text == BTN_RENDER_DOC)
async def render_doc_start(m: Message, state: FSMContext):
    if not _FITZ_OK:
        await m.answer("Конвертация недоступна: библиотека PyMuPDF не установлена.")
        return
    if shutil.which("libreoffice") is None and shutil.which("soffice") is None:
        await m.answer("Конвертация DOC/DOCX недоступна: LibreOffice не установлен в окружении.")
        return
    logger.info("render: start DOC session user=%s", m.from_user.id)
    await reset_render_state(state)
    await state.set_state(RenderSession.waiting_file)
    await state.update_data(
        render_format="docx",
        render_items=[],
        render_card_mid=None,
        render_choose_mid=None,
        render_channels={},
        render_index=0,
        render_wm_text=None,
    )
    await m.answer(
        "Пришлите Word-файл (.doc или .docx) как документ — мы превратим его в PNG 300 DPI. "
        "⚠️ Ограничение Telegram: размер файла до 20 МБ."
    )


@router.message(F.text == BTN_RENDER_XLSX)
async def render_xlsx_start(m: Message, state: FSMContext):
    if not _FITZ_OK:
        await m.answer("Конвертация недоступна: библиотека PyMuPDF не установлена.")
        return
    if not _OPENPYXL_OK:
        await m.answer("Обработка Excel недоступна: пакет openpyxl не установлен.")
        return
    if shutil.which("libreoffice") is None and shutil.which("soffice") is None:
        await m.answer("Конвертация Excel недоступна: LibreOffice не установлен в окружении.")
        return
    logger.info("render: start XLSX session user=%s", m.from_user.id)
    await reset_render_state(state)
    await state.set_state(RenderSession.waiting_file)
    await state.update_data(
        render_format="xlsx",
        render_items=[],
        render_card_mid=None,
        render_choose_mid=None,
        render_channels={},
        render_index=0,
        render_wm_text=None,
    )
    await m.answer(
        "Пришлите Excel-файл (.xlsx, .xls) как документ. Мы определим таблицы на листах, покажем превью и подготовим PNG 300 DPI. "
        "Если листов несколько, отметьте нужные таблицы и нажмите «Продолжить»."
    )


@router.message(RenderSession.waiting_file, F.document)
async def render_file_receive(m: Message, state: FSMContext):
    data = await state.get_data()
    render_format = (data.get("render_format") or "pdf").lower()
    doc = m.document
    file_size = int(doc.file_size or 0)
    if file_size > MAX_FILE_SIZE:
        await m.answer("Файл превышает ограничение Telegram (20 МБ). Сожмите документ перед отправкой.")
        return

    if render_format == "pdf":
        default_name = "document.pdf"
    elif render_format == "docx":
        default_name = "document.docx"
    elif render_format == "png":
        default_name = "image.png"
    else:
        default_name = "spreadsheet.xlsx"
    filename = (doc.file_name or default_name).strip()
    ext = Path(filename).suffix.lower()

    if render_format == "pdf" and ext != ".pdf":
        await m.answer("Пожалуйста, отправьте файл с расширением .pdf.")
        return
    if render_format == "docx" and ext not in (".doc", ".docx"):
        await m.answer("Пожалуйста, отправьте файл Word (.doc или .docx).")
        return
    if render_format == "png" and ext != ".png":
        await m.answer("Пожалуйста, отправьте PNG-файл.")
        return
    if render_format == "xlsx" and ext not in (".xlsx", ".xls", ".xlsm", ".ods", ".fods"):
        await m.answer("Пожалуйста, отправьте Excel-файл (.xlsx, .xls, .ods).")
        return
    if render_format == "xlsx" and not _OPENPYXL_OK:
        await m.answer("Обработка Excel недоступна: пакет openpyxl не установлен на сервере.")
        return

    if render_format == "xlsx":
        status_text = "Анализирую файл..."
    elif render_format == "docx":
        status_text = "Анализирую документ..."
    elif render_format == "png":
        status_text = "Обрабатываю PNG..."
    else:
        status_text = "Конвертируем файл..."
    status_msg = await m.answer(status_text)
    try:
        logger.info("render: downloading file format=%s name=%s size=%s", render_format, filename, file_size)
        file = await m.bot.get_file(doc.file_id)
        downloaded = await m.bot.download_file(file.file_path)
        if isinstance(downloaded, (bytes, bytearray, memoryview)):
            blob = bytes(downloaded)
        elif hasattr(downloaded, "getvalue"):
            blob = downloaded.getvalue()
        elif hasattr(downloaded, "read"):
            blob = downloaded.read()
        else:
            blob = bytes(downloaded)
        if not isinstance(blob, (bytes, bytearray)):
            if hasattr(blob, "getvalue"):
                blob = blob.getvalue()
            elif hasattr(blob, "read"):
                blob = blob.read()
            elif hasattr(blob, "tobytes"):
                blob = blob.tobytes()
            else:
                blob = bytes(blob)
        logger.info("render: downloaded bytes=%s", len(blob))

        storage_key: str | None = None
        prefix = SOURCE_PREFIXES.get(render_format, "file")
        try:
            storage_key = await store_blob(prefix, blob)
            preview_result = await _fetch_preview_from_worker(
                render_format,
                filename,
                storage_key=storage_key,
                blob=blob,
            )
        except Exception as exc:
            if storage_key:
                await delete_blob(storage_key)
            logger.exception("render: preview failed format=%s name=%s", render_format, filename)
            await status_msg.edit_text(_format_error("Не удалось подготовить страницы", exc))
            return

        pages_raw = list(preview_result.get("pages") or [])
        analysis = preview_result.get("analysis") or {}
        if not pages_raw:
            if storage_key:
                await delete_blob(storage_key)
            raise RuntimeError("Не удалось подготовить страницы для предпросмотра.")

        logger.info("render: preview ready format=%s name=%s bytes=%s -> pages=%s", render_format, filename, len(blob), len(pages_raw))

        new_pages: List[Dict[str, Any]] = []
        for entry in pages_raw:
            preview_key = entry.get("preview_key")
            if not preview_key:
                logger.warning("render: preview key missing for %s", filename)
                continue
            try:
                preview_bytes = await load_blob(preview_key, delete=True)
            except Exception:
                logger.warning("render: failed to load preview page for %s (key=%s)", filename, preview_key)
                continue

            page_info: Dict[str, Any] = {
                "filename": entry.get("filename") or filename,
                "original_bytes": None,
                "watermarked_bytes": None,
                "preview_original_bytes": preview_bytes,
                "preview_watermarked_bytes": None,
                "selected": True,
            }
            fullres_key = entry.get("fullres_key")
            if fullres_key:
                page_info["fullres_key"] = fullres_key

            for key in ("sheet_name", "table_range", "sheet_index", "sheets_total", "table_index", "tables_in_sheet", "base_name", "display_name", "page_index", "pages_total"):
                if key in entry:
                    page_info[key] = entry[key]

            if render_format == "png":
                page_info["source_key"] = storage_key
                if fullres_key is None and storage_key:
                    page_info["fullres_key"] = storage_key
            new_pages.append(page_info)

        new_item: Dict[str, Any] = {
            "source": filename,
            "format": render_format,
            "source_key": storage_key,
            "pages": new_pages,
        }
        if render_format == "xlsx":
            new_item["sheets_total"] = analysis.get("sheets_total")

        lock = _get_render_lock(m.from_user.id)
        async with lock:
            latest = await state.get_data()
            items: List[Dict[str, Any]] = list(latest.get("render_items") or [])
            items.append(new_item)
            storage_key = None

            wm_text = latest.get("render_wm_text")
            if wm_text:
                try:
                    await _apply_watermark_to_items([new_item], wm_text)
                except Exception as e:
                    logger.exception("render: watermark failed format=%s name=%s", render_format, filename)
                    await m.answer(_format_error("Не удалось применить водяной знак", e))

            await state.set_state(RenderSession.idle)
            await state.update_data(render_items=items)

            try:
                await status_msg.delete()
            except Exception:
                pass

            try:
                await _update_render_card(m.bot, m.chat.id, state, focus="last")
            except Exception as e:
                logger.exception("render: failed to update preview format=%s name=%s", render_format, filename)
                await m.answer(f"Не удалось сформировать превью: {e}")
    except Exception as e:
        try:
            if locals().get("storage_key"):
                await delete_blob(storage_key)
        except Exception:
            pass
        logger.exception("render: failed to process file format=%s name=%s size=%s", render_format, filename, file_size)
        message = _format_error("Не удалось обработать файл", e)
        try:
            await status_msg.edit_text(message)
        except Exception:
            await m.answer(message)

@router.message(RenderSession.waiting_file)
async def render_file_waiting_other(m: Message, state: FSMContext):
    data = await state.get_data()
    render_format = (data.get("render_format") or "pdf").lower()
    if render_format == "docx":
        text = "Сейчас ожидаем Word-файл (.doc или .docx) как документ (до 20 МБ)."
    elif render_format == "xlsx":
        text = "Сейчас ожидаем Excel-файл (.xlsx, .xls, .ods) как документ (до 20 МБ)."
    elif render_format == "png":
        text = "Сейчас ожидаем PNG-файл как документ (до 20 МБ)."
    else:
        text = "Сейчас ожидаем PDF-файл как документ (до 20 МБ)."
    await m.answer(text)



@router.callback_query(F.data == "render:add")
async def render_pdf_add(cq: CallbackQuery, state: FSMContext):
    await state.set_state(RenderSession.waiting_file)
    data = await state.get_data()
    render_format = (data.get("render_format") or "pdf").lower()
    if render_format == "docx":
        text = "Пришлите следующий Word-файл (.doc или .docx) как документ (до 20 МБ)."
    elif render_format == "xlsx":
        text = "Пришлите следующий Excel-файл (.xlsx, .xls, .ods) как документ (до 20 МБ)."
    elif render_format == "png":
        text = "Пришлите следующий PNG-файл как документ (до 20 МБ)."
    else:
        text = "Пришлите следующий PDF-файл как документ (до 20 МБ)."
    await cq.message.answer(text)
    await cq.answer()


@router.callback_query(F.data == "render:prev")
async def render_pdf_prev(cq: CallbackQuery, state: FSMContext):
    lock = _get_render_lock(cq.from_user.id)
    async with lock:
        data = await state.get_data()
        items = data.get("render_items") or []
        total = len(_flatten_pages(items))
        if total > 1:
            index = (data.get("render_index", 0) - 1) % total
            await state.update_data(render_index=index)
            await _update_render_card(cq.bot, cq.message.chat.id, state)
    await cq.answer()


@router.callback_query(F.data == "render:next")
async def render_pdf_next(cq: CallbackQuery, state: FSMContext):
    lock = _get_render_lock(cq.from_user.id)
    async with lock:
        data = await state.get_data()
        items = data.get("render_items") or []
        total = len(_flatten_pages(items))
        if total > 1:
            index = (data.get("render_index", 0) + 1) % total
            await state.update_data(render_index=index)
            await _update_render_card(cq.bot, cq.message.chat.id, state)
    await cq.answer()


@router.callback_query(F.data == "render:toggle")
async def render_toggle_selection(cq: CallbackQuery, state: FSMContext):
    response_text: str | None = None
    show_alert = False
    lock = _get_render_lock(cq.from_user.id)
    async with lock:
        data = await state.get_data()
        render_format = (data.get("render_format") or "pdf").lower()
        if render_format in {"xlsx", "docx", "pdf", "png"}:
            items: List[Dict[str, Any]] = list(data.get("render_items") or [])
            if not items:
                response_text = "Нет элементов для выбора."
                show_alert = True
            else:
                flat = _flatten_pages(items)
                if not flat:
                    response_text = "Нет элементов для выбора."
                    show_alert = True
                else:
                    index = data.get("render_index", 0)
                    index = max(0, min(index, len(flat) - 1))
                    item_idx, page_idx = flat[index]
                    page = items[item_idx]["pages"][page_idx]
                    page["selected"] = not page.get("selected", True)
                    await state.update_data(render_items=items)
                    await _update_render_card(cq.bot, cq.message.chat.id, state)
                    response_text = "Добавлено в выборку." if page["selected"] else "Исключено из выборки."
        # если формат не поддерживает выбор, просто вернём пустой ответ
    if response_text is None:
        await cq.answer()
    else:
        await cq.answer(response_text, show_alert=show_alert)


@router.callback_query(F.data == "render:wm:set")
async def render_pdf_wm_request(cq: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    items = data.get("render_items") or []
    if not items:
        await cq.answer("Сначала добавьте файлы для конвертации.", show_alert=True)
        return
    if not _PIL_OK:
        await cq.answer("Водяной знак недоступен: Pillow не установлен.", show_alert=True)
        return
    await state.set_state(RenderSession.waiting_wm_text)
    await cq.message.answer("Напишите текст для водяного знака.")
    await cq.answer()


@router.message(RenderSession.waiting_wm_text)
async def render_pdf_wm_text(m: Message, state: FSMContext):
    text = (m.text or "").strip()
    if not text:
        await m.answer("Текст пустой. Укажите текст для водяного знака.")
        return
    lock = _get_render_lock(m.from_user.id)
    async with lock:
        data = await state.get_data()
        items: List[Dict[str, Any]] = list(data.get("render_items") or [])
        if not items:
            await m.answer("Нет файлов для применения водяного знака.")
            await state.set_state(RenderSession.idle)
            return
        try:
            await _apply_watermark_to_items(items, text)
        except Exception as e:
            await m.answer(_format_error("Не удалось применить водяной знак", e))
            await state.set_state(RenderSession.idle)
            return
        await state.update_data(render_items=items, render_wm_text=text)
        await state.set_state(RenderSession.idle)
        await m.answer("Водяной знак добавлен.")
        await _update_render_card(m.bot, m.chat.id, state)


@router.callback_query(F.data == "render:wm:clear")
async def render_pdf_wm_clear(cq: CallbackQuery, state: FSMContext):
    response_text = "Водяной знак удалён."
    show_alert = False
    lock = _get_render_lock(cq.from_user.id)
    async with lock:
        data = await state.get_data()
        items: List[Dict[str, Any]] = list(data.get("render_items") or [])
        if not items:
            response_text = "Нет файлов для очистки."
            show_alert = True
        else:
            _clear_watermarks(items)
            await state.update_data(render_items=items, render_wm_text=None)
            await _update_render_card(cq.bot, cq.message.chat.id, state)
    await cq.answer(response_text, show_alert=show_alert)


@router.callback_query(F.data == "render:cancel")
async def render_pdf_cancel(cq: CallbackQuery, state: FSMContext):
    lock = _get_render_lock(cq.from_user.id)
    async with lock:
        await _clear_render_context(cq.bot, cq.message.chat.id, state)
        sent = await cq.message.answer("Выберите формат для конвертации:", reply_markup=build_render_menu_keyboard())
        await state.update_data(menu_mid=sent.message_id)
    await cq.answer("Отменено.")


@router.callback_query(F.data == "render:upload")
async def render_pdf_upload(cq: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    items = data.get("render_items") or []
    if not items:
        await cq.answer("Сначала добавьте файлы для конвертации.", show_alert=True)
        return
    render_format = (data.get("render_format") or "pdf").lower()
    if render_format in {"xlsx", "docx", "pdf", "png"}:
        has_selected = any(page.get("selected", True) for item in items for page in item["pages"])
        if not has_selected:
            if render_format == "xlsx":
                message = "Выберите хотя бы одну таблицу."
            elif render_format == "png":
                message = "Выберите хотя бы один файл."
            else:
                message = "Выберите хотя бы одну страницу."
            await cq.answer(message, show_alert=True)
            return

    contractor_id = str(cq.from_user.id)
    channels = await _fetch_recent_channels(contractor_id)
    if not channels:
        await cq.answer()
        await cq.message.answer("Каналы пока не найдены. Создайте канал через «🆕 Новый канал».")
        return

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=ch["title"], callback_data=f"render:ch:{ch['channel_id']}")]
            for ch in channels
        ]
    )
    sent = await cq.message.answer("Выберите канал для загрузки:", reply_markup=kb)
    await state.update_data(
        render_choose_mid=sent.message_id,
        render_channels={str(ch["channel_id"]): ch["title"] for ch in channels},
    )
    await cq.answer()


@router.callback_query(F.data.startswith("render:ch:"))
async def render_pdf_upload_to_channel(cq: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    items: List[Dict[str, Any]] = list(data.get("render_items") or [])
    if not items:
        await cq.answer("Нет подготовленных файлов.", show_alert=True)
        return
    render_format = (data.get("render_format") or "pdf").lower()
    if render_format in {"xlsx", "docx", "pdf", "png"}:
        selected_count = sum(1 for item in items for page in item["pages"] if page.get("selected", True))
        if selected_count == 0:
            if render_format == "xlsx":
                message = "Нет выбранных таблиц для отправки."
            elif render_format == "png":
                message = "Нет выбранных файлов для отправки."
            else:
                message = "Нет выбранных страниц для отправки."
            await cq.answer(message, show_alert=True)
            return

    _, _, channel_id_str = cq.data.partition("render:ch:")
    try:
        channel_id = int(channel_id_str)
    except ValueError:
        await cq.answer("Неверный канал.", show_alert=True)
        return

    channels_map = data.get("render_channels") or {}
    channel_title = channels_map.get(channel_id_str, "канал")
    wm_text: str | None = data.get("render_wm_text")
    if wm_text:
        try:
            await _ensure_watermark_for_all(items, wm_text)
        except Exception as e:
            await cq.answer(f"Не удалось подготовить водяной знак: {e}", show_alert=True)
            return

    await cq.answer("Готовим файлы к загрузке…")

    use_worker = _USE_CELERY_PUBLISH
    celery_app = get_celery() if use_worker else None
    publish_queue = os.getenv("CELERY_PUBLISH_QUEUE", "publish")
    pdf_queue = os.getenv("CELERY_PDF_QUEUE", "pdf")
    office_queue = os.getenv("CELERY_OFFICE_QUEUE", "office")

    for item in items:
        item_format = str(item.get("format") or render_format).lower()
        pages = list(item.get("pages") or [])
        selected_pages = []
        for idx, page in enumerate(pages, start=1):
            if item_format in {"xlsx", "docx", "pdf", "png"} and not page.get("selected", True):
                continue
            selected_pages.append((int(page.get("page_index") or idx), page))
        if not selected_pages:
            continue

        if not use_worker:
            fullres_cleanup: Set[str] = set()
            for _, page in selected_pages:
                await _ensure_page_original_bytes(page)
                payload = None
                if wm_text:
                    payload = page.get("watermarked_bytes") or page.get("preview_watermarked_bytes")
                else:
                    payload = page.get("original_bytes") or page.get("preview_original_bytes")
                if not payload:
                    continue
                filename = page.get("filename") or "smeta.png"
                try:
                    await cq.bot.send_document(
                        chat_id=channel_id,
                        document=BufferedInputFile(payload, filename=filename),
                        protect_content=True,
                    )
                    key_for_cleanup = page.get("fullres_key")
                    if key_for_cleanup:
                        fullres_cleanup.add(key_for_cleanup)
                except Exception as e:
                    await cq.message.answer(_format_error(f"�� 㤠���� ��ࠢ��� {filename}", e))
            if fullres_cleanup:
                await delete_many(fullres_cleanup)
            await delete_blob(item.get("source_key"))
            continue
        if celery_app is None:
            continue

        png_pages: List[Tuple[int, Dict[str, Any], str]] = []
        missing_pages: List[Tuple[int, Dict[str, Any]]] = []
        for page_index, page in selected_pages:
            fullres_key = page.get("fullres_key")
            if fullres_key:
                png_pages.append((page_index, page, fullres_key))
            else:
                missing_pages.append((page_index, page))

        for _, page, fullres_key in png_pages:
            celery_app.send_task(
                "tasks.render.process_and_publish_png",
                kwargs={
                    "chat_id": channel_id,
                    "png_key": fullres_key,
                    "watermark_text": wm_text,
                    "filename": page.get("filename") or "smeta.png",
                    "apply_watermark": bool(wm_text),
                },
                queue=publish_queue,
            )

        if not missing_pages:
            continue

        if item_format == "png":
            for _, page in missing_pages:
                fallback_key = page.get("source_key") or item.get("source_key")
                if fallback_key:
                    celery_app.send_task(
                        "tasks.render.process_and_publish_png",
                        kwargs={
                            "chat_id": channel_id,
                            "png_key": fallback_key,
                            "watermark_text": wm_text,
                            "filename": page.get("filename") or "smeta.png",
                            "apply_watermark": bool(wm_text),
                        },
                        queue=publish_queue,
                    )
                    continue
                await _ensure_page_original_bytes(page)
                page_bytes = page.get("original_bytes") or page.get("preview_original_bytes")
                if not page_bytes:
                    continue
                encoded = base64.b64encode(page_bytes).decode("ascii")
                celery_app.send_task(
                    "tasks.render.process_and_publish_png",
                    kwargs={
                        "chat_id": channel_id,
                        "png_b64": encoded,
                        "watermark_text": wm_text,
                        "filename": page.get("filename") or "smeta.png",
                        "apply_watermark": bool(wm_text),
                    },
                    queue=publish_queue,
                )
            continue

        source_key = item.get("source_key")
        if not source_key:
            await cq.message.answer(f"�� ������ �������� ���� ��� {item.get('source') or '���������'}.")
            continue

        page_numbers = [idx for idx, _ in missing_pages]
        task_kwargs = {
            "chat_id": channel_id,
            "watermark_text": wm_text,
            "filename": item.get("source") or "document",
            "page_indices": page_numbers,
        }

        if item_format == "pdf":
            task_kwargs["pdf_key"] = source_key
            celery_app.send_task(
                "tasks.render.process_and_publish_pdf",
                kwargs=task_kwargs,
                queue=pdf_queue,
            )
        elif item_format == "docx":
            task_kwargs["doc_key"] = source_key
            celery_app.send_task(
                "tasks.render.process_and_publish_doc",
                kwargs=task_kwargs,
                queue=office_queue,
            )
        elif item_format == "xlsx":
            task_kwargs["excel_key"] = source_key
            celery_app.send_task(
                "tasks.render.process_and_publish_excel",
                kwargs=task_kwargs,
                queue=office_queue,
            )
        else:
            await cq.message.answer(f"������ {item_format} ���� �� ��������� ��� ������� ����������.")
    choose_mid = data.get("render_choose_mid")
    if choose_mid:
        try:
            await cq.bot.delete_message(chat_id=cq.message.chat.id, message_id=choose_mid)
        except Exception:
            pass
    card_mid = data.get("render_card_mid")
    if card_mid:
        try:
            await cq.bot.delete_message(chat_id=cq.message.chat.id, message_id=card_mid)
        except Exception:
            pass

    await reset_render_state(state)
    if use_worker:
        confirmation = (
            f"PNG файлы поставлены в очередь для публикации в канал «{channel_title}». "
            "Готовые изображения появятся в канале в течение нескольких секунд. "
            "Перейдите в раздел «🔗 Мои ссылки» для получения уникальной ссылки на канал. "
            "В разделе «📢 Мои каналы» можете управлять уже созданными каналами."
        )
    else:
        confirmation = (
            f"PNG файлы загружены в канал «{channel_title}». "
            "Перейдите в раздел «🔗 Мои ссылки» для получения уникальной ссылки на канал. "
            "В разделе «📢 Мои каналы» можете управлять уже созданными каналами."
        )
    sent = await cq.message.answer(confirmation, reply_markup=build_render_menu_keyboard())
    await state.update_data(menu_mid=sent.message_id)

