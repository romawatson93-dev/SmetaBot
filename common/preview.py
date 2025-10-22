from __future__ import annotations

import base64
import io
import os
import re
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Tuple

try:
    import fitz  # type: ignore

    _FITZ_OK = True
except Exception:  # pragma: no cover - optional dependency
    fitz = None  # type: ignore
    _FITZ_OK = False

try:
    from PIL import Image

    _PIL_OK = True
except Exception:  # pragma: no cover - optional dependency
    Image = None  # type: ignore
    _PIL_OK = False

try:
    import openpyxl  # type: ignore
    from openpyxl.utils import get_column_letter, range_boundaries  # type: ignore
    from openpyxl.worksheet.page import PageMargins  # type: ignore
    from openpyxl.worksheet.properties import PageSetupProperties  # type: ignore

    _OPENPYXL_OK = True
except Exception:  # pragma: no cover - optional dependency
    openpyxl = None  # type: ignore
    get_column_letter = None  # type: ignore
    range_boundaries = None  # type: ignore
    PageMargins = None  # type: ignore
    PageSetupProperties = None  # type: ignore
    _OPENPYXL_OK = False

_SANITIZE_RE = re.compile(r"[^A-Za-z0-9._-]+")


class PreviewError(RuntimeError):
    """Raised when preview generation fails."""


def _sanitize_basename(filename: str, default: str = "document") -> str:
    base = Path(filename).stem or default
    sanitized = _SANITIZE_RE.sub("_", base)
    return sanitized[:48] or default


def _make_preview(png_bytes: bytes, max_dim: int = 1600) -> bytes:
    if not _PIL_OK or Image is None:  # pragma: no cover - fallback path
        return png_bytes
    with Image.open(io.BytesIO(png_bytes)) as img:
        img = img.convert("RGB")
        img.thumbnail((max_dim, max_dim), Image.LANCZOS)
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=85, optimize=True)
        return out.getvalue()


def _convert_pdf(pdf_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    if not _FITZ_OK or fitz is None:
        raise PreviewError("PyMuPDF (fitz) недоступен в окружении превью.")
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")  # type: ignore[call-arg]
    pages: List[Dict[str, Any]] = []
    try:
        total = doc.page_count
        base = Path(filename).stem or "page"
        for idx, page in enumerate(doc, start=1):
            pix = page.get_pixmap(dpi=300, alpha=False)
            out_name = f"{base}-{idx:02}.png" if total > 1 else f"{base}.png"
            pages.append(
                {
                    "filename": out_name,
                    "content": pix.tobytes("png"),
                    "page_index": idx,
                    "pages_total": total,
                }
            )
    finally:
        doc.close()
    return pages


def _wrap_png_as_pages(png_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    base_name = Path(filename).stem or "image"
    return [
        {
            "filename": f"{base_name}.png",
            "content": png_bytes,
            "page_index": 1,
            "pages_total": 1,
        }
    ]


def _convert_doc_to_pdf_bytes(doc_bytes: bytes, suffix: str) -> bytes:
    suffix = suffix.lower()
    allowed = {".doc", ".docx", ".xls", ".xlsx", ".xlsm", ".ods", ".fods"}
    if suffix not in allowed:
        raise PreviewError(f"Формат {suffix or 'неизвестно'} не поддерживается для конвертации в PDF.")
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        src_path = tmpdir_path / f"source{suffix}"
        src_path.write_bytes(doc_bytes)
        binary = shutil.which("libreoffice") or shutil.which("soffice")
        if not binary:
            raise PreviewError("LibreOffice не найден в окружении превью.")
        cmd = [
            binary,
            "--headless",
            "--nologo",
            "--nodefault",
            "--nofirststartwizard",
            "--norestore",
            "--nolockcheck",
            "--convert-to",
            "pdf",
            "--outdir",
            str(tmpdir_path),
            str(src_path),
        ]
        env = os.environ.copy()
        env.setdefault("HOME", str(tmpdir_path))
        env.setdefault("TMPDIR", str(tmpdir_path))
        proc = subprocess.run(
            cmd,
            capture_output=True,
            timeout=240,
            cwd=tmpdir_path,
            env=env,
        )
        if proc.returncode != 0:
            raise PreviewError(
                "LibreOffice не смог обработать документ: "
                f"{proc.stderr.decode(errors='ignore') or proc.stdout.decode(errors='ignore')}"
            )
        pdf_path = src_path.with_suffix(".pdf")
        if not pdf_path.exists():
            candidates = list(tmpdir_path.glob("*.pdf"))
            if not candidates:
                raise PreviewError("LibreOffice не создал PDF-файл.")
            pdf_path = candidates[0]
        return pdf_path.read_bytes()


def _convert_doc_to_png(doc_bytes: bytes, suffix: str, filename: str) -> List[Dict[str, Any]]:
    pdf_bytes = _convert_doc_to_pdf_bytes(doc_bytes, suffix)
    return _convert_pdf(pdf_bytes, filename)


def _range_to_a1(bounds: Tuple[int, int, int, int]) -> str:
    if get_column_letter is None:
        raise PreviewError("openpyxl недоступен в окружении превью.")
    min_col, min_row, max_col, max_row = bounds
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _prepare_excel_bytes_for_openpyxl(excel_bytes: bytes, suffix: str) -> Tuple[bytes, str]:
    if suffix == ".xls":
        with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
            tmp.write(excel_bytes)
            tmp.flush()
            try:
                binary = shutil.which("libreoffice") or shutil.which("soffice")
                if not binary:
                    raise PreviewError("LibreOffice не найден, XLS нельзя обработать.")
                with tempfile.TemporaryDirectory() as tmpdir:
                    out_path = Path(tmpdir) / "converted.xlsx"
                    cmd = [
                        binary,
                        "--headless",
                        "--convert-to",
                        "xlsx",
                        str(Path(tmp.name)),
                        "--outdir",
                        tmpdir,
                    ]
                    proc = subprocess.run(cmd, capture_output=True, timeout=240)
                    if proc.returncode != 0 or not out_path.exists():
                        raise PreviewError("Не удалось конвертировать XLS в XLSX через LibreOffice.")
                    return out_path.read_bytes(), ".xlsx"
            finally:
                os.unlink(tmp.name)
    return excel_bytes, suffix


def _worksheet_detect_tables(ws) -> List[Tuple[int, int, int, int]]:
    if range_boundaries is None or get_column_letter is None:
        raise PreviewError("openpyxl недоступен в окружении превью.")
    regions: List[Tuple[int, int, int, int]] = []
    for table in getattr(ws, "tables", {}).values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        regions.append((min_col, min_row, max_col, max_row))
    if regions:
        return regions
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row and max_col:
        regions.append((1, 1, max_col, max_row))
    return regions


def _export_excel_region_to_pdf(excel_bytes: bytes, suffix: str, sheet_name: str, bounds: Tuple[int, int, int, int]) -> bytes:
    if not _OPENPYXL_OK or openpyxl is None:
        raise PreviewError("openpyxl недоступен, невозможно обработать Excel.")
    prepared_bytes, prepared_suffix = _prepare_excel_bytes_for_openpyxl(excel_bytes, suffix)
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_path = tmp_path / f"source{prepared_suffix}"
        src_path.write_bytes(prepared_bytes)
        wb = openpyxl.load_workbook(src_path, data_only=False)  # type: ignore[arg-type]
        try:
            if sheet_name not in wb.sheetnames:
                raise PreviewError(f"Лист {sheet_name!r} не найден в книге.")
            for ws in wb.worksheets:
                ws.sheet_state = "hidden" if ws.title != sheet_name else "visible"
            ws = wb[sheet_name]
            area = _range_to_a1(bounds)
            ws.print_area = area
            if PageSetupProperties is not None:
                ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)  # type: ignore[attr-defined]
            ws.page_setup.fitToWidth = 1  # type: ignore[attr-defined]
            ws.page_setup.fitToHeight = 1  # type: ignore[attr-defined]
            orientation = "landscape" if (bounds[3] - bounds[2]) > (bounds[1] - bounds[0]) else "portrait"
            ws.page_setup.orientation = orientation  # type: ignore[attr-defined]
            if PageMargins is not None:
                ws.page_margins = PageMargins(
                    left=0.25,
                    right=0.25,
                    top=0.3,
                    bottom=0.3,
                    header=0.1,
                    footer=0.1,
                )
            wb.active = wb.sheetnames.index(sheet_name)
            wb.save(src_path)
        finally:
            wb.close()
        return _convert_doc_to_pdf_bytes(src_path.read_bytes(), ".xlsx")


def _extract_excel_tables(excel_bytes: bytes, filename: str) -> Dict[str, Any]:
    if not _OPENPYXL_OK or openpyxl is None:
        raise PreviewError("openpyxl недоступен, не удаётся обработать Excel.")
    suffix = Path(filename).suffix or ".xlsx"
    prepared_bytes, prepared_suffix = _prepare_excel_bytes_for_openpyxl(excel_bytes, suffix)
    wb = openpyxl.load_workbook(io.BytesIO(prepared_bytes), data_only=True)  # type: ignore[arg-type]
    try:
        total_sheets = len(wb.worksheets)
        orig_base = (os.path.splitext(filename)[0] or "document").strip()
        base_name = _sanitize_basename(filename)
        tables: List[Dict[str, Any]] = []
        for sheet_index, ws in enumerate(wb.worksheets):
            regions = _worksheet_detect_tables(ws)
            if not regions:
                continue
            tables_in_sheet = len(regions)
            for table_idx, bounds in enumerate(regions, start=1):
                pdf_bytes = _export_excel_region_to_pdf(prepared_bytes, prepared_suffix, ws.title, bounds)
                png_pages = _convert_pdf(
                    pdf_bytes,
                    f"{base_name}-sheet{sheet_index + 1}-tbl{table_idx}.pdf",
                )
                if not png_pages:
                    continue
                first = png_pages[0]
                display_name = (orig_base or "document") + ".png"
                tables.append(
                    {
                        "filename": display_name,
                        "content": first["content"],
                        "sheet_name": ws.title,
                        "table_range": _range_to_a1(bounds),
                        "sheet_index": sheet_index,
                        "sheets_total": total_sheets,
                        "table_index": table_idx,
                        "tables_in_sheet": tables_in_sheet,
                        "base_name": base_name,
                        "display_name": display_name,
                        "page_index": first.get("page_index", len(tables) + 1),
                        "pages_total": first.get("pages_total"),
                    }
                )
        return {"pages": tables, "sheets_total": total_sheets}
    finally:
        wb.close()


def generate_preview(file_bytes: bytes, filename: str, render_format: str) -> Dict[str, Any]:
    fmt = (render_format or '').strip().lower()
    if fmt not in {"pdf", "docx", "xlsx", "png"}:
        raise PreviewError(f"Неизвестный формат превью: {render_format}")

    analysis: Dict[str, Any] = {}
    if fmt == "pdf":
        pages_raw = _convert_pdf(file_bytes, filename)
    elif fmt == "docx":
        suffix = Path(filename).suffix or ".docx"
        pages_raw = _convert_doc_to_png(file_bytes, suffix, filename)
    elif fmt == "png":
        pages_raw = _wrap_png_as_pages(file_bytes, filename)
    else:  # fmt == "xlsx"
        analysis = _extract_excel_tables(file_bytes, filename)
        pages_raw = list(analysis.get("pages") or [])

    if not pages_raw:
        raise PreviewError("Не удалось подготовить страницы для превью.")

    pages: List[Dict[str, Any]] = []
    for entry in pages_raw:
        png_bytes = entry.get("content")
        if not png_bytes:
            continue
        preview = _make_preview(png_bytes)
        page_info: Dict[str, Any] = {
            "filename": entry.get("filename") or filename,
            "preview_bytes": preview,
            "fullres_bytes": png_bytes,
        }
        for key in (
            "sheet_name",
            "table_range",
            "sheet_index",
            "sheets_total",
            "table_index",
            "tables_in_sheet",
            "base_name",
            "display_name",
            "page_index",
            "pages_total",
        ):
            if key in entry:
                page_info[key] = entry[key]
        pages.append(page_info)

    return {"pages": pages, "analysis": {k: v for k, v in analysis.items() if k != "pages"}}


__all__ = ["generate_preview", "PreviewError"]

